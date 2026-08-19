"""图片 OCR -> Excel 转换器。

支持的输入格式: ``.png`` / ``.jpg`` / ``.jpeg`` / ``.bmp`` / ``.tiff``
输出格式: ``.xlsx``

主要能力:
    - 使用 Tesseract OCR 引擎（通过 ``pytesseract``）识别图片中的文字。
    - 默认语言 ``chi_sim+eng``（简体中文 + 英文）。
    - 通过 ``image_to_data`` 提取每个词的位置、置信度等 layout 信息。
    - 同一 ``block_num / par_num / line_num`` 的词合并为一行，
      并按 ``left`` 升序写入 Excel 单元格，从而在视觉上近似还原原图布局。
    - 多张图片可合并为同一 xlsx，每个图片对应一个 sheet。

异常处理:
    - 缺少 tesseract 二进制 / 缺少 pytesseract 库 / 指定语言包未安装
      均抛出 ``ConversionError``，并附带安装提示。
    - 在 ``convert_many`` 中，单张图片 OCR 失败会记录 warning 并继续，
      最终生成一个空 sheet（保持 sheet 数量与输入一致），由调用方决定是否重试。
"""

from __future__ import annotations

from pathlib import Path
from typing import ClassVar, Iterable, List, Tuple

from PIL import Image, UnidentifiedImageError

try:  # pytesseract 是可选依赖；缺失时延迟到 convert 阶段再报错
    import pytesseract  # type: ignore[import-untyped]
    from pytesseract import TesseractNotFoundError  # type: ignore[import-untyped]
    from pytesseract import TesseractError  # type: ignore[import-untyped]
    from pytesseract import Output  # type: ignore[import-untyped]
    _PYTESSERACT_AVAILABLE = True
except Exception:  # noqa: BLE001 - 延迟到运行时再判定
    pytesseract = None  # type: ignore[assignment]
    TesseractNotFoundError = Exception  # type: ignore[assignment,misc]
    TesseractError = Exception  # type: ignore[assignment,misc]
    Output = None  # type: ignore[assignment]
    _PYTESSERACT_AVAILABLE = False

try:
    from openpyxl import Workbook
    _OPENPYXL_AVAILABLE = True
except Exception:  # noqa: BLE001
    Workbook = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False

from ..core.base import BaseConverter, ConversionError, PathLike
from ..core.logger import get_logger
from ..core.registry import Registry
from ..utils.paths import ensure_dir, unique_output_path


class OcrConverter(BaseConverter):
    """图片 OCR 转 Excel 转换器。

    Examples:
        单文件::

            converter = OcrConverter()
            converter.convert("scan.png", "result.xlsx")

        批量（每个图一个 sheet）::

            converter.convert_many(
                ["page1.png", "page2.png"],
                "all_pages.xlsx",
            )
    """

    name: ClassVar[str] = "OcrConverter"
    supported_pairs: ClassVar[Tuple[Tuple[str, str], ...]] = (
        (".png", ".xlsx"),
        (".jpg", ".xlsx"),
        (".jpeg", ".xlsx"),
        (".bmp", ".xlsx"),
        (".tiff", ".xlsx"),
    )

    _log = get_logger("converters.ocr")

    def __init__(
        self,
        *,
        lang: str = "chi_sim+eng",
        min_confidence: int = 50,
        **kwargs: object,
    ) -> None:
        """初始化 OCR 转换器。

        Args:
            lang: Tesseract 语言标识。默认 ``chi_sim+eng``。
            min_confidence: 丢弃置信度低于此阈值的词（0-100）。默认 50。
            **kwargs: 预留扩展参数。
        """
        if not lang or not isinstance(lang, str):
            raise ValueError(f"lang 必须是非空字符串，收到: {lang!r}")
        if not (0 <= int(min_confidence) <= 100):
            raise ValueError(
                f"min_confidence 必须在 [0, 100]，收到: {min_confidence}"
            )
        self.lang: str = lang
        self.min_confidence: int = int(min_confidence)

    # ------------------------------------------------------------------ #
    # 公开 API
    # ------------------------------------------------------------------ #
    def convert(
        self,
        source: PathLike,
        target: PathLike,
        **kwargs: object,
    ) -> Path:
        """将单张图片 OCR 结果写入 Excel（单 sheet）。

        Args:
            source: 源图片路径。
            target: 输出 xlsx 路径。
            **kwargs: 支持 ``lang``、``min_confidence``、``overwrite``。

        Returns:
            实际写入的 xlsx 路径。

        Raises:
            ConversionError: 环境/参数/文件错误。
        """
        self._ensure_dependencies()
        src, dst = self._resolve_paths(source, target)
        self._check_pair_supported(self.supported_pairs, src.suffix, dst.suffix)

        lang, min_conf, overwrite = self._extract_kwargs(kwargs)

        ensure_dir(dst.parent)
        final_dst = dst if overwrite else unique_output_path(dst)

        workbook = Workbook()
        try:
            workbook.remove(workbook.active)
            sheet_name = self._safe_sheet_name(src.stem)
            ws = workbook.create_sheet(title=sheet_name)

            words = self._ocr_file(src, lang=lang, min_conf=min_conf)
            self._layout_to_sheet(ws, words)

            workbook.save(final_dst)
        except OSError as exc:
            raise ConversionError(f"写入 xlsx 失败: {final_dst} ({exc})") from exc
        finally:
            # openpyxl 的 Workbook 没有显式 close；保留以防后续扩展
            pass

        self._log.info(
            "OCR 转 Excel 完成: %s -> %s (词数=%d)", src, final_dst, self._last_word_count
        )
        return final_dst

    def convert_many(
        self,
        sources: Iterable[PathLike],
        target: PathLike,
        **kwargs: object,
    ) -> Path:
        """将多张图片的 OCR 结果合并到同一 xlsx，每个图片一个 sheet。

        单张图片失败会被记录为 warning 并继续（写入一个空 sheet），
        整体流程不会被中断。

        Args:
            sources: 图片路径列表。
            target: 输出 xlsx 路径。
            **kwargs: 支持 ``lang``、``min_confidence``、``overwrite``。

        Returns:
            实际写入的 xlsx 路径。
        """
        self._ensure_dependencies()

        src_list: List[Path] = []
        for item in sources:
            p = Path(item).expanduser()
            if not p.exists():
                raise ConversionError(f"源文件不存在: {p}")
            src_list.append(p)

        if not src_list:
            raise ConversionError("convert_many 至少需要一张图片")

        dst = Path(target).expanduser()
        if not dst.suffix:
            raise ConversionError(f"目标文件必须包含扩展名 .xlsx: {dst}")
        self._check_pair_supported(self.supported_pairs, src_list[0].suffix, dst.suffix)

        lang, min_conf, overwrite = self._extract_kwargs(kwargs)

        ensure_dir(dst.parent)
        final_dst = dst if overwrite else unique_output_path(dst)

        workbook = Workbook()
        try:
            workbook.remove(workbook.active)
            used_names: set[str] = set()
            for src in src_list:
                sheet_name = self._safe_sheet_name(src.stem, used=used_names)
                ws = workbook.create_sheet(title=sheet_name)
                try:
                    words = self._ocr_file(src, lang=lang, min_conf=min_conf)
                    self._layout_to_sheet(ws, words)
                except ConversionError as exc:
                    self._log.warning(
                        "单张图片 OCR 失败，跳过内容写入 (%s): %s", src, exc
                    )

            workbook.save(final_dst)
        except OSError as exc:
            raise ConversionError(f"写入 xlsx 失败: {final_dst} ({exc})") from exc

        self._log.info(
            "OCR 批量转 Excel 完成: %d 张图 -> %s",
            len(src_list),
            final_dst,
        )
        return final_dst

    # ------------------------------------------------------------------ #
    # OCR 核心
    # ------------------------------------------------------------------ #
    def _ocr_file(
        self,
        src: Path,
        *,
        lang: str,
        min_conf: int,
    ) -> List[dict]:
        """读取图片文件并提取词级 OCR 数据。"""
        try:
            image = Image.open(src)
        except (FileNotFoundError, UnidentifiedImageError) as exc:
            raise ConversionError(f"无法读取图片: {src} ({exc})") from exc

        try:
            return self._ocr_image(image, lang=lang, min_conf=min_conf)
        finally:
            try:
                image.close()
            except Exception:  # noqa: BLE001
                pass

    def _ocr_image(
        self,
        img: Image.Image,
        *,
        lang: str | None = None,
        min_conf: int | None = None,
    ) -> List[dict]:
        """对已加载的 :class:`PIL.Image.Image` 执行 OCR。

        返回词级 dict 列表，每个 dict 包含:
            ``text``、``conf``、``left``、``top``、``width``、``height``、
            ``block``、``par``、``line``。
        """
        self._ensure_dependencies()

        use_lang = lang or self.lang
        use_conf = self.min_confidence if min_conf is None else int(min_conf)
        if not (0 <= use_conf <= 100):
            raise ConversionError(
                f"min_confidence 必须在 [0, 100]，收到: {use_conf}"
            )

        try:
            data = pytesseract.image_to_data(  # type: ignore[union-attr]
                img,
                lang=use_lang,
                output_type=Output.DICT,  # type: ignore[union-attr]
            )
        except TesseractNotFoundError as exc:  # type: ignore[misc]
            raise ConversionError(
                "未找到 tesseract 二进制，请先安装 Tesseract OCR。"
                "Ubuntu/Debian: apt-get install tesseract-ocr tesseract-ocr-chi-sim；"
                "macOS: brew install tesseract tesseract-lang；"
                "Windows: https://github.com/UB-Mannheim/tesseract/wiki"
            ) from exc
        except TesseractError as exc:  # type: ignore[misc]
            # 常见原因：语言包未安装
            msg = str(exc)
            if "could not create TXT output file" in msg or "language" in msg.lower():
                raise ConversionError(
                    f"Tesseract 语言包加载失败 (lang={use_lang}): {msg}。"
                    "请确认已安装对应的 tesseract 语言包，例如"
                    " tesseract-ocr-chi-sim。"
                ) from exc
            raise ConversionError(f"OCR 失败: {msg}") from exc

        words: List[dict] = []
        n = len(data.get("text", []))
        for i in range(n):
            text = (data["text"][i] or "").strip()
            if not text:
                continue
            try:
                conf = int(float(data["conf"][i]))
            except (TypeError, ValueError):
                continue
            if conf < use_conf:
                continue
            try:
                left = int(data["left"][i])
                top = int(data["top"][i])
                width = int(data["width"][i])
                height = int(data["height"][i])
            except (TypeError, ValueError):
                # 坐标缺失的词直接丢弃
                continue
            words.append({
                "text": text,
                "conf": conf,
                "left": left,
                "top": top,
                "width": width,
                "height": height,
                "block": int(data["block_num"][i]),
                "par": int(data["par_num"][i]),
                "line": int(data["line_num"][i]),
            })

        self._last_word_count = len(words)
        return words

    def _layout_to_sheet(self, ws, words: List[dict]) -> None:
        """把词级 OCR 结果按行写入 Excel 单元格。

        行分组策略: 相同 ``(block, par, line)`` 视为同一行；
        之后按 ``top`` 升序排版到 Excel 的 1, 2, 3... 行；
        行内按 ``left`` 升序填入 1, 2, 3... 列。
        """
        if not words:
            return

        rows: dict[Tuple[int, int, int], List[dict]] = {}
        for w in words:
            key = (w["block"], w["par"], w["line"])
            rows.setdefault(key, []).append(w)

        # 按最小 top 升序（视觉上的"由上到下"）
        sorted_rows = sorted(rows.values(), key=lambda r: min(x["top"] for x in r))
        for r_idx, row_words in enumerate(sorted_rows, start=1):
            row_words.sort(key=lambda w: w["left"])
            for c_idx, w in enumerate(row_words, start=1):
                ws.cell(row=r_idx, column=c_idx, value=w["text"])

    # ------------------------------------------------------------------ #
    # 工具方法
    # ------------------------------------------------------------------ #
    def _ensure_dependencies(self) -> None:
        """检查 ``pytesseract`` 和 ``openpyxl`` 是否可用。"""
        if not _PYTESSERACT_AVAILABLE:
            raise ConversionError(
                "缺少依赖 pytesseract，请先安装: pip install pytesseract"
            )
        if not _OPENPYXL_AVAILABLE:
            raise ConversionError(
                "缺少依赖 openpyxl，请先安装: pip install openpyxl"
            )
        # 提前探测 tesseract 二进制是否存在，给出最清晰的错误
        try:
            pytesseract.get_tesseract_version()  # type: ignore[union-attr]
        except TesseractNotFoundError as exc:  # type: ignore[misc]
            raise ConversionError(
                "未找到 tesseract 二进制，请先安装 Tesseract OCR。"
                "Ubuntu/Debian: apt-get install tesseract-ocr tesseract-ocr-chi-sim；"
                "macOS: brew install tesseract tesseract-lang；"
                "Windows: https://github.com/UB-Mannheim/tesseract/wiki"
            ) from exc

    def _extract_kwargs(
        self,
        kwargs: dict,
    ) -> Tuple[str, int, bool]:
        """从 kwargs 提取 lang / min_confidence / overwrite，统一做校验。"""
        lang = kwargs.pop("lang", self.lang)
        if not lang or not isinstance(lang, str):
            raise ConversionError(f"lang 必须是非空字符串，收到: {lang!r}")
        min_conf_raw = kwargs.pop("min_confidence", self.min_confidence)
        try:
            min_conf = int(min_conf_raw)  # type: ignore[arg-type]
        except (TypeError, ValueError) as exc:
            raise ConversionError(
                f"min_confidence 必须是整数，收到: {min_conf_raw!r}"
            ) from exc
        if not (0 <= min_conf <= 100):
            raise ConversionError(
                f"min_confidence 必须在 [0, 100]，收到: {min_conf}"
            )
        overwrite = bool(kwargs.pop("overwrite", False))
        return lang, min_conf, overwrite

    @staticmethod
    def _safe_sheet_name(name: str, used: set[str] | None = None) -> str:
        """生成合法的 Excel sheet 名（最长 31 字符，去掉非法字符，避免冲突）。"""
        illegal = set('[]:*?/\\')
        cleaned = "".join("_" if ch in illegal else ch for ch in name).strip()
        if not cleaned:
            cleaned = "sheet"
        cleaned = cleaned[:31] or "sheet"
        if used is None:
            return cleaned
        # 在批量场景下避免重名
        original = cleaned
        i = 2
        while cleaned in used:
            suffix = f"_{i}"
            cleaned = (original[: 31 - len(suffix)] + suffix) if len(original) + len(suffix) > 31 else original + suffix
            i += 1
        used.add(cleaned)
        return cleaned

    # 内部状态：最近一次 OCR 提取到的词数（仅供日志使用，不作为对外 API）
    _last_word_count: int = 0


# ---------------------------------------------------------------------------
# 模块加载时自动注册到全局路由。
# ---------------------------------------------------------------------------
Registry.register(OcrConverter())
