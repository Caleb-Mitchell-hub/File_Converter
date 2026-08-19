"""基于 OpenCV 几何检测的图片表格 OCR → Excel 转换器。

针对**带完整横竖边框**的标准表格设计：

1. 用形态学 + 投影检测所有横线、竖线，得到 ``(n_rows+1, n_cols+1)`` 个网格交点。
2. 把网格交点构造为 ``cells[i][j]`` 矩形列表。
3. 推断合并单元格：
   - 横向合并：相邻若干 ``cells[i][j..k]`` 之间没有内部分隔线（检测该列带位置没有竖线）。
   - 纵向合并：相邻若干 ``cells[i..l][j]`` 之间没有内部分隔线（该行带位置没有横线）。
4. 按格子用 Tesseract 识别文字。
5. 写入 openpyxl，保留合并。

如果**完全检测不到网格**（无线表格），抛 ``ConversionError`` 让上层选择走 VL 模型。
"""

from __future__ import annotations

import logging
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, ClassVar, Iterable, List, Tuple

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image

from ..core.base import BaseConverter, ConversionError, PathLike
from ..core.registry import Registry
from ..utils.paths import ensure_dir, unique_output_path

_log = logging.getLogger("doc_converter.opencv_ocr")

# 默认参数
MIN_TABLE_AREA = 0.20        # 检测到的表格覆盖图面积比（小于则判为无线表格）
H_PROJECT_RATIO = 0.30       # 横线投影阈值 = max * ratio
V_PROJECT_RATIO = 0.30
MIN_ROWS = 2                 # 至少 2 行
MIN_COLS = 2                 # 至少 2 列


# ---------------------------------------------------------------------- #
# 数据结构
# ---------------------------------------------------------------------- #
@dataclass
class CellRect:
    """一个网格单元格的矩形坐标（0-based）。"""
    r: int
    c: int
    x0: int
    y0: int
    x1: int
    y1: int

    @property
    def width(self) -> int:
        return self.x1 - self.x0

    @property
    def height(self) -> int:
        return self.y1 - self.y0


@dataclass
class TableGrid:
    """完整的表格结构。"""
    cells: List[List[CellRect]]   # cells[r][c] 给出矩形坐标
    n_rows: int
    n_cols: int


# ---------------------------------------------------------------------- #
# 转换器
# ---------------------------------------------------------------------- #
class OpenCvOcrConverter(BaseConverter):
    """OpenCV 几何检测 + Tesseract OCR。"""

    name: ClassVar[str] = "OpenCvOcrConverter"
    supported_pairs: ClassVar[Tuple[Tuple[str, str], ...]] = (
        (".png", ".xlsx"),
        (".jpg", ".xlsx"),
        (".jpeg", ".xlsx"),
        (".bmp", ".xlsx"),
        (".tiff", ".xlsx"),
        (".webp", ".xlsx"),
    )

    def convert(
        self,
        source: PathLike,
        target: PathLike,
        **kwargs: Any,
    ) -> Path:
        src, dst = self._resolve_paths(source, target)
        self._check_pair_supported(self.supported_pairs, src.suffix, dst.suffix)
        overwrite = bool(kwargs.get("overwrite", False))

        ensure_dir(dst.parent)
        final_dst = dst if overwrite else unique_output_path(dst)

        # 1) 检测网格 + 线集合（OpenCV 几何）
        img = _read_image(str(src))
        grid = detect_grid(img)
        lines = detect_lines(img)

        # 2) 推断合并（OpenCV）
        opencv_merges = infer_merges(grid, lines)

        # 3) 调 Qwen-VL 拿文字（更准）
        use_qwen = kwargs.get("use_qwen_vl", True)
        if use_qwen:
            try:
                hybrid = _ocr_with_qwen_vl_full(grid, src_path=str(src))
                # 用 Qwen-VL 的细网格 + Qwen-VL 自己的 merges
                # （OpenCV merges 在 Qwen-VL 列划分上不对齐，舍弃）
                final_cell_text = hybrid.cell_text
                final_n_rows = hybrid.n_rows
                final_n_cols = hybrid.n_cols
                final_merges = hybrid.merges
            except ConversionError as exc:
                _log.warning("Qwen-VL OCR 失败，回退到 Tesseract: %s", exc)
                final_cell_text = ocr_cells(grid, src_path=str(src),
                                            lang=kwargs.get("lang", "chi_sim+eng"))
                final_n_rows = grid.n_rows
                final_n_cols = grid.n_cols
                final_merges = [(r1, c1, r2, c2, None) for r1, c1, r2, c2 in opencv_merges]
        else:
            final_cell_text = ocr_cells(grid, src_path=str(src),
                                        lang=kwargs.get("lang", "chi_sim+eng"))
            final_n_rows = grid.n_rows
            final_n_cols = grid.n_cols
            final_merges = [(r1, c1, r2, c2, None) for r1, c1, r2, c2 in opencv_merges]

        # 4) 渲染：用最终决定的网格 + merges
        wb = Workbook()
        try:
            wb.remove(wb.active)
            ws = wb.create_sheet(title=_safe_sheet_name(src.stem))
            _render_hybrid(ws, final_n_rows, final_n_cols, final_cell_text, final_merges)
            wb.save(final_dst)
        finally:
            pass

        _log.info(
            "OpenCV+Qwen-VL 完成: %s -> %s (rows=%d, cols=%d, merges=%d)",
            src, final_dst, final_n_rows, final_n_cols, len(final_merges),
        )
        return final_dst

    def convert_many(
        self,
        sources: Iterable[PathLike],
        target: PathLike,
        **kwargs: Any,
    ) -> Path:
        src_list: List[Path] = []
        for item in sources:
            p = Path(item).expanduser()
            if not p.exists():
                raise ConversionError(f"源文件不存在: {p}")
            src_list.append(p)
        if not src_list:
            raise ConversionError("convert_many 至少需要一张图片")

        dst = Path(target).expanduser()
        if not dst.suffix:
            raise ConversionError(f"目标文件必须包含扩展名 .xlsx: {dst}")
        self._check_pair_supported(self.supported_pairs, src_list[0].suffix, dst.suffix)
        overwrite = bool(kwargs.get("overwrite", False))

        ensure_dir(dst.parent)
        final_dst = dst if overwrite else unique_output_path(dst)

        wb = Workbook()
        try:
            wb.remove(wb.active)
            used_names: set[str] = set()
            for src in src_list:
                ws = wb.create_sheet(title=_safe_sheet_name(src.stem, used=used_names))
                try:
                    img = _read_image(str(src))
                    grid = detect_grid(img)
                    lines = detect_lines(img)
                    opencv_merges = infer_merges(grid, lines)
                    use_qwen = kwargs.get("use_qwen_vl", True)
                    if use_qwen:
                        try:
                            hybrid = _ocr_with_qwen_vl_full(grid, src_path=str(src))
                            final_cell_text = hybrid.cell_text
                            final_n_rows = hybrid.n_rows
                            final_n_cols = hybrid.n_cols
                            final_merges = hybrid.merges
                        except ConversionError as exc:
                            _log.warning("Qwen-VL OCR 失败，回退到 Tesseract: %s", exc)
                            final_cell_text = ocr_cells(grid, src_path=str(src),
                                                        lang=kwargs.get("lang", "chi_sim+eng"))
                            final_n_rows = grid.n_rows
                            final_n_cols = grid.n_cols
                            final_merges = [(r1, c1, r2, c2, None) for r1, c1, r2, c2 in opencv_merges]
                    else:
                        final_cell_text = ocr_cells(grid, src_path=str(src),
                                                    lang=kwargs.get("lang", "chi_sim+eng"))
                        final_n_rows = grid.n_rows
                        final_n_cols = grid.n_cols
                        final_merges = [(r1, c1, r2, c2, None) for r1, c1, r2, c2 in opencv_merges]
                    _render_hybrid(ws, final_n_rows, final_n_cols, final_cell_text, final_merges)
                except ConversionError as exc:
                    _log.warning("OpenCV OCR 单图失败，留空 sheet (%s): %s", src, exc)
            wb.save(final_dst)
        except OSError as exc:
            raise ConversionError(f"写入 xlsx 失败: {final_dst} ({exc})") from exc

        return final_dst


# ---------------------------------------------------------------------- #
# 图像读取
# ---------------------------------------------------------------------- #
def _read_image(path: str) -> np.ndarray:
    """读取图片为单通道灰度图（uint8）。"""
    img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        raise ConversionError(f"无法读取图片: {path}")
    return img


# ---------------------------------------------------------------------- #
# 1. 检测网格
# ---------------------------------------------------------------------- #
@dataclass
class GridLine:
    """检测到的一条线 + 它的"实际可见长度"（像素）。"""
    coord: int          # 横线 y / 竖线 x
    length_px: int      # 实际跨越的像素长度


@dataclass
class GridLines:
    """detect_lines 的返回：每条线的坐标 + 实际长度。"""
    h_lines: List[GridLine]
    v_lines: List[GridLine]

    @property
    def h_lines_y(self) -> List[int]:
        return [l.coord for l in self.h_lines]

    @property
    def v_lines_x(self) -> List[int]:
        return [l.coord for l in self.v_lines]


def detect_grid(img: np.ndarray) -> TableGrid:
    """检测图片中的表格网格。

    复用 ``detect_lines`` 拿到的线集合，保证 grid 和 merge 推断使用同一组线。
    """
    lines = detect_lines(img)
    h_lines_y = lines.h_lines_y
    v_lines_x = lines.v_lines_x

    # 验证
    if len(h_lines_y) < MIN_ROWS + 1:
        raise ConversionError(
            f"OpenCV 网格检测失败: 仅检测到 {len(h_lines_y)} 条横线，"
            f"至少需要 {MIN_ROWS + 1} 条。"
            "可能是无线表格，请改用 Qwen-VL 路径。"
        )
    if len(v_lines_x) < MIN_COLS + 1:
        raise ConversionError(
            f"OpenCV 网格检测失败: 仅检测到 {len(v_lines_x)} 条竖线，"
            f"至少需要 {MIN_COLS + 1} 条。"
            "可能是无线表格，请改用 Qwen-VL 路径。"
        )

    n_rows = len(h_lines_y) - 1
    n_cols = len(v_lines_x) - 1
    cells: List[List[CellRect]] = []
    for r in range(n_rows):
        row_cells: List[CellRect] = []
        for c in range(n_cols):
            row_cells.append(CellRect(
                r=r, c=c,
                x0=v_lines_x[c], y0=h_lines_y[r],
                x1=v_lines_x[c + 1], y1=h_lines_y[r + 1],
            ))
        cells.append(row_cells)

    return TableGrid(cells=cells, n_rows=n_rows, n_cols=n_cols)


# ---------------------------------------------------------------------- #
# 3b. 混合 OCR：Qwen-VL 提供文字，OpenCV 提供合并区域
# ---------------------------------------------------------------------- #
@dataclass
class HybridGrid:
    """混合网格：以 Qwen-VL 的细网格为基础，应用 Qwen-VL 的合并区域。"""
    cell_text: List[List[str]]
    n_rows: int
    n_cols: int
    merges: List[Tuple[int, int, int, int, str | None]]   # (r1,c1,r2,c2,value)


def _ocr_with_qwen_vl_full(grid: TableGrid, *, src_path: str) -> HybridGrid:
    """调 Qwen-VL 拿文字 + 用 OpenCV grid 做合并推断。

    返回 ``HybridGrid``：
    - ``n_rows / n_cols``：以 Qwen-VL 为准（更细）。
    - ``cell_text[r][c]``：Qwen-VL 识别的文字。
    - ``merges``：OpenCV 检测的合并区域（应用到 Qwen-VL 网格上）。

    由于 OpenCV 的合并区域基于"宽列"判定，可能不完全适用于 Qwen-VL 的"细列"，
    这里用启发式映射：把 OpenCV 的 merge 按列宽比例放大到 Qwen-VL 网格。
    """
    try:
        from app.config import get_settings  # type: ignore[import-not-found]
        s = get_settings()
        cfg_kwargs = {
            "api_key": s.qwen_api_key,
            "base_url": s.qwen_base_url,
            "model": s.qwen_model,
            "timeout": s.qwen_timeout,
        }
    except Exception as exc:  # noqa: BLE001
        raise ConversionError(
            "Qwen-VL OCR 不可用：无法读取 app.config"
        ) from exc

    from .qwen_vl_client import QwenVlConfig, QwenVlError, chat_with_image
    if not cfg_kwargs.get("api_key"):
        raise ConversionError("Qwen-VL OCR 不可用：未配置 QWEN_API_KEY")

    cfg = QwenVlConfig(**cfg_kwargs)
    try:
        result = chat_with_image(cfg=cfg, image_path=src_path)
    except QwenVlError as exc:
        raise ConversionError(f"Qwen-VL OCR 失败: {exc}") from exc

    header = result.get("header") or []
    rows = result.get("rows") or []
    qwen_merges = result.get("merges") or []

    # Qwen-VL 网格：以 header+rows 拼接得到原始行数，最大列宽为 n_cols
    qwen_rows: List[List[str]] = []
    for r in header + rows:
        qwen_rows.append([str(c) if c is not None else "" for c in r])
    qwen_n_cols = max((len(r) for r in qwen_rows), default=0)
    qwen_n_rows = len(qwen_rows)

    # Qwen-VL 给出的 merges，保留 value 字段供渲染时写入 anchor 单元格
    from .qwen_ocr_converter import _norm_cell  # type: ignore[import-not-found]
    qwen_merges_tuples: List[Tuple[int, int, int, int, str | None]] = []
    for m in qwen_merges:
        try:
            r1, c1, r2, c2 = int(m["r1"]), int(m["c1"]), int(m["r2"]), int(m["c2"])
        except Exception:
            continue
        if r2 < r1 or c2 < c1:
            continue
        value = _norm_cell(m.get("value"))
        qwen_merges_tuples.append((r1, c1, r2, c2, value))

    return HybridGrid(
        cell_text=qwen_rows,
        n_rows=qwen_n_rows,
        n_cols=qwen_n_cols,
        merges=qwen_merges_tuples,
    )


def detect_lines(img: np.ndarray) -> GridLines:
    """检测横竖线 + 每条线的实际可见长度。

    实现：
        1) 形态学 OPEN 找到候选线位置（坐标 + 一组像素）；
        2) 对每个候选坐标，在原图 binary 上量最长连续黑像素段 = 该线长度。
    """
    H, W = img.shape[:2]
    blur = cv2.GaussianBlur(img, (3, 3), 0)
    _, binary = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    # 注：binary 中线条 = 255（白），背景 = 0（黑）

    h_kernel_len = max(20, W // 6)
    v_kernel_len = max(20, H // 6)
    h_morph = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (h_kernel_len, 1)))
    v_morph = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_kernel_len)))

    h_proj = h_morph.sum(axis=1).astype(np.float32) / 255.0
    v_proj = v_morph.sum(axis=0).astype(np.float32) / 255.0

    def _find_peaks(arr, ratio):
        thresh = arr.max() * ratio if arr.max() > 0 else 0
        return [int(i) for i, v in enumerate(arr) if v >= thresh]

    def _merge_close(peaks, min_gap):
        if not peaks:
            return []
        peaks = sorted(peaks)
        merged = [peaks[0]]
        for p in peaks[1:]:
            if p - merged[-1] >= min_gap:
                merged.append(p)
        return merged

    def _longest_run_x(y: int) -> int:
        """在 binary[y, :] 上找最长连续 255 的长度（横线的长度）。"""
        row = binary[y, :]
        max_run = run = 0
        for v in row:
            if v > 0:
                run += 1
                if run > max_run:
                    max_run = run
            else:
                run = 0
        return max_run

    def _longest_run_y(x: int) -> int:
        col = binary[:, x]
        max_run = run = 0
        for v in col:
            if v > 0:
                run += 1
                if run > max_run:
                    max_run = run
            else:
                run = 0
        return max_run

    h_peaks = _find_peaks(h_proj, H_PROJECT_RATIO)
    v_peaks = _find_peaks(v_proj, V_PROJECT_RATIO)
    h_lines_y = _merge_close(h_peaks, max(3, H // 60))
    v_lines_x = _merge_close(v_peaks, max(3, W // 60))

    # 计算每条线的实际长度（跳过噪点 — 太短的忽略）
    min_line_len = max(8, min(H, W) // 30)

    # 字符笔画 vs 真竖线判别：
    # 字符竖笔画（如"批""一"等）经过 v_morph OPEN 后，v_proj 仍然 ≥ 0.95 * max；
    # 真表格竖线 v_proj 通常 ≤ 0.90 * max。
    # 横线方向没有这种字符笔画干扰，统一保留。
    if v_proj.max() > 0:
        v_char_threshold = v_proj.max() * 0.92
    else:
        v_char_threshold = 1e9

    h_lines: List[GridLine] = []
    for y in h_lines_y:
        L = _longest_run_x(y)
        if L >= min_line_len:
            h_lines.append(GridLine(coord=y, length_px=L))
    v_lines: List[GridLine] = []
    for x in v_lines_x:
        if v_proj[x] >= v_char_threshold:
            # 大概率是字符笔画的"假竖线"，跳过
            continue
        L = _longest_run_y(x)
        if L >= min_line_len:
            v_lines.append(GridLine(coord=x, length_px=L))

    return GridLines(h_lines=h_lines, v_lines=v_lines)


# ---------------------------------------------------------------------- #
# 2. 推断合并单元格
# ---------------------------------------------------------------------- #
def infer_merges(
    grid: TableGrid,
    lines: GridLines,
    tolerance: int = 6,
) -> List[Tuple[int, int, int, int]]:
    """根据 detect_lines 推断合并区域。

    判定规则：**只把"贯穿整张图"的线视为真分隔线**。
    - 真表格竖线：长度 ≥ 90% 的图高度；
    - 真表格横线：长度 ≥ 90% 的图宽度。
    - 部分长度（合并区边框、字符笔画）的线 → 视为不分割 → 合并。

    这种启发式简单但对带完整边框的标准表格非常有效：
    - 合并区通常有"边框"但不会延伸过相邻区域；
    - 真行/列分隔线一定贯穿整图。
    """
    # 满长阈值：图中横线最长长度、竖线最长长度
    if lines.h_lines:
        max_h = max(l.length_px for l in lines.h_lines)
        full_h_threshold = max_h * 0.90
    else:
        full_h_threshold = 0

    if lines.v_lines:
        max_v = max(l.length_px for l in lines.v_lines)
        full_v_threshold = max_v * 0.90
    else:
        full_v_threshold = 0

    def _find_nearest(target: int, candidates: List[GridLine], tol: int) -> GridLine | None:
        best: GridLine | None = None
        best_dist = tol + 1
        for c in candidates:
            d = abs(c.coord - target)
            if d <= tol and d < best_dist:
                best = c
                best_dist = d
        return best

    def _has_horizontal_separator(c1: CellRect, c2: CellRect) -> bool:
        """c1 在左, c2 在右：mid_x 处竖线是不是"贯穿整图"的真分隔线。"""
        if c1.x1 != c2.x0:
            return False
        line = _find_nearest(c1.x1, lines.v_lines, tolerance)
        if line is None:
            return False
        return line.length_px >= full_v_threshold

    def _has_vertical_separator(c1: CellRect, c2: CellRect) -> bool:
        if c1.y1 != c2.y0:
            return False
        line = _find_nearest(c1.y1, lines.h_lines, tolerance)
        if line is None:
            return False
        return line.length_px >= full_h_threshold

    parent: List[int] = []
    def find(i: int) -> int:
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i
    def union(i: int, j: int) -> None:
        ri, rj = find(i), find(j)
        if ri != rj:
            parent[rj] = ri

    idx = [[r * grid.n_cols + c for c in range(grid.n_cols)] for r in range(grid.n_rows)]
    for r in range(grid.n_rows):
        for c in range(grid.n_cols):
            parent.append(idx[r][c])

    for r in range(grid.n_rows):
        for c in range(grid.n_cols - 1):
            if not _has_horizontal_separator(grid.cells[r][c], grid.cells[r][c + 1]):
                union(idx[r][c], idx[r][c + 1])
    for r in range(grid.n_rows - 1):
        for c in range(grid.n_cols):
            if not _has_vertical_separator(grid.cells[r][c], grid.cells[r + 1][c]):
                union(idx[r][c], idx[r + 1][c])

    groups: dict[int, List[Tuple[int, int]]] = {}
    for r in range(grid.n_rows):
        for c in range(grid.n_cols):
            root = find(idx[r][c])
            groups.setdefault(root, []).append((r, c))

    merges: List[Tuple[int, int, int, int]] = []
    for cells_in_group in groups.values():
        rs = [rc[0] for rc in cells_in_group]
        cs = [rc[1] for rc in cells_in_group]
        r1, r2 = min(rs), max(rs)
        c1, c2 = min(cs), max(cs)
        if r1 == r2 and c1 == c2:
            continue
        merges.append((r1, c1, r2, c2))
    return merges


# ---------------------------------------------------------------------- #
# 4. 渲染（接受任意 rows/cols + merges + text，不依赖 OpenCV grid）
# ---------------------------------------------------------------------- #
def _render_hybrid(
    ws,
    n_rows: int,
    n_cols: int,
    cell_text: List[List[str]],
    merges: List[Tuple[int, int, int, int, str | None]],
) -> None:
    """通用渲染：写文字 + 应用合并。

    合并逻辑与 ``QwenOcrConverter._render_table`` 一致：
    1) 先 merges → 建立覆盖集 + 调 openpyxl merge_cells
    2) 写非合并区格子文字
    3) 写合并区 anchor 格子的 value（优先用 merge 自带的 value）
    """
    # 规范化 cell_text：补齐每行列数 + None -> "" 归一化
    from .qwen_ocr_converter import _norm_cell  # type: ignore[import-not-found]

    padded: List[List[str | None]] = []
    for r in range(n_rows):
        if r < len(cell_text):
            row = cell_text[r]
        else:
            row = []
        padded.append([_norm_cell(row[c]) if c < len(row) else None for c in range(n_cols)])

    # 1) 建立"合并覆盖集合" + 真正合并单元格
    merge_region: set[Tuple[int, int]] = set()
    for r1, c1, r2, c2, _v in merges:
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                merge_region.add((rr, cc))
        ws.merge_cells(
            start_row=r1 + 1, start_column=c1 + 1,
            end_row=r2 + 1, end_column=c2 + 1,
        )

    # 2) 写非合并区格子（跳过所有 merge_region 格子，避免触碰 MergedCell）
    for r in range(n_rows):
        for c in range(n_cols):
            if (r, c) in merge_region:
                continue
            val = padded[r][c]
            if val is not None:
                ws.cell(row=r + 1, column=c + 1, value=val)

    # 3) 写合并区 anchor 格子的 value（优先用 merge 自带的 value，
    #    若 merge 无 value 则回退到 cell_text 对应格子的内容）
    for r1, c1, r2, c2, v in merges:
        if v is not None:
            ws.cell(row=r1 + 1, column=c1 + 1, value=v)
        else:
            # 回退：用 cell_text 中 anchor 格子的内容
            fallback = padded[r1][c1] if r1 < len(padded) and c1 < len(padded[r1]) else None
            if fallback is not None:
                ws.cell(row=r1 + 1, column=c1 + 1, value=fallback)

    # 4) 列宽
    from openpyxl.utils import get_column_letter
    for c in range(1, max(n_cols, 1) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14


# ---------------------------------------------------------------------- #
# 3. 按格子 OCR
# ---------------------------------------------------------------------- #
def _ensure_tesseract_cmd() -> None:
    """显式设置 tesseract 可执行路径 + tessdata 目录。

    Windows 默认安装路径；其它平台请通过环境变量 TESSERACT_CMD 覆盖。
    tessdata 目录通过环境变量 ``TESSDATA_PREFIX`` 显式注入，
    否则 subprocess 不会继承当前 shell 的环境变量。
    """
    try:
        import pytesseract
    except Exception:
        return  # 留到调用时再报错
    import os
    cmd = os.environ.get("TESSERACT_CMD")
    if not cmd:
        candidates = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        ]
        for c in candidates:
            if os.path.exists(c):
                cmd = c
                break
    if cmd:
        pytesseract.pytesseract.tesseract_cmd = cmd

    # tessdata 路径：通过 monkey-patch subprocess 注入环境变量
    tessdata = os.environ.get("TESSDATA_PREFIX")
    if not tessdata:
        # 默认用户目录下的 tessdata
        candidates = [
            os.path.join(os.path.expanduser("~"), "tessdata"),
            r"C:\Program Files\Tesseract-OCR\tessdata",
            r"C:\Program Files (x86)\Tesseract-OCR\tessdata",
        ]
        for c in candidates:
            if os.path.isdir(c):
                tessdata = c
                break
    if tessdata:
        # pytesseract 内部用 subprocess.run(env=os.environ)，所以修改 os.environ 即可
        os.environ["TESSDATA_PREFIX"] = tessdata


def ocr_cells(grid: TableGrid, *, src_path: str, lang: str = "chi_sim+eng") -> List[List[str]]:
    """对每个 cell 调用 Tesseract 识别文字。

    为了让 OCR 准确，先把每个 cell 裁切出来（加 padding），放大 2x，再调 tesseract。
    Tesseract 期望原图（白底黑字），不要做反相。
    返回二维数组 ``text[r][c]``。
    """
    try:
        import pytesseract
    except Exception as exc:  # noqa: BLE001
        raise ConversionError(
            "缺少依赖 pytesseract，请先安装: pip install pytesseract"
        ) from exc

    _ensure_tesseract_cmd()

    img = Image.open(src_path).convert("L")  # 灰度
    try:
        n_rows, n_cols = grid.n_rows, grid.n_cols
        result: List[List[str]] = []
        for r in range(n_rows):
            row_text: List[str] = []
            for c in range(n_cols):
                rect = grid.cells[r][c]
                pad = 4
                x0 = max(0, rect.x0 - pad)
                y0 = max(0, rect.y0 - pad)
                x1 = min(img.width, rect.x1 + pad)
                y1 = min(img.height, rect.y1 + pad)
                if x1 - x0 < 4 or y1 - y0 < 4:
                    row_text.append("")
                    continue
                crop = img.crop((x0, y0, x1, y1))
                # 放大 2x 提升小字识别
                big = crop.resize((crop.width * 2, crop.height * 2), Image.LANCZOS)
                # PSM 6: 视为统一文本块（自动判断单/多行）
                try:
                    txt = pytesseract.image_to_string(
                        big, lang=lang, config="--psm 6"
                    ).strip()
                except Exception:
                    txt = ""
                row_text.append(txt)
            result.append(row_text)
        return result
    finally:
        try:
            img.close()
        except Exception:
            pass


# ---------------------------------------------------------------------- #
# 辅助
# ---------------------------------------------------------------------- #
def _safe_sheet_name(name: str, used: set[str] | None = None) -> str:
    illegal = set("[]:*?/\\")
    cleaned = "".join("_" if ch in illegal else ch for ch in name).strip()
    if not cleaned:
        cleaned = "sheet"
    cleaned = cleaned[:31] or "sheet"
    if used is None:
        return cleaned
    original = cleaned
    i = 2
    while cleaned in used:
        suffix = f"_{i}"
        cleaned = (original[: 31 - len(suffix)] + suffix) if len(original) + len(suffix) > 31 else original + suffix
        i += 1
    used.add(cleaned)
    return cleaned


# 模块加载时自动注册到全局路由
Registry.register(OpenCvOcrConverter())