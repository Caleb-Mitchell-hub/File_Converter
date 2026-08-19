"""基于 Qwen-VL 大模型的图片 OCR → Excel 转换器。

与 :class:`OcrConverter`（本地 Tesseract）相比：
- 优先调用 DashScope OpenAI 兼容接口的 Qwen-VL-Plus。
- 支持识别**合并单元格**（横跨多列/多行的表头与数据）。
- 仅当 API key 未配置时直接抛 ``ConversionError``，
  避免悄悄走本地 Tesseract 后端掩盖配置错误。

输入/输出格式与 ``OcrConverter`` 完全一致，便于服务层透明切换。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, ClassVar, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..core.base import BaseConverter, ConversionError, PathLike
from ..core.registry import Registry
from ..utils.paths import ensure_dir, unique_output_path
from .qwen_vl_client import QwenVlConfig, QwenVlError, chat_with_image

_log = logging.getLogger("doc_converter.qwen_ocr")


class QwenOcrConverter(BaseConverter):
    """使用 Qwen-VL 把图片中的表格识别为带合并单元格的 Excel。"""

    name: ClassVar[str] = "QwenOcrConverter"
    supported_pairs: ClassVar[Tuple[Tuple[str, str], ...]] = (
        (".png", ".xlsx"),
        (".jpg", ".xlsx"),
        (".jpeg", ".xlsx"),
        (".bmp", ".xlsx"),
        (".tiff", ".xlsx"),
        (".webp", ".xlsx"),
    )

    def __init__(self, cfg: QwenVlConfig | None = None, **kwargs: Any) -> None:
        self.cfg = cfg

    # ------------------------------------------------------------------ #
    # 公开 API
    # ------------------------------------------------------------------ #
    def convert(
        self,
        source: PathLike,
        target: PathLike,
        **kwargs: Any,
    ) -> Path:
        src, dst = self._resolve_paths(source, target)
        self._check_pair_supported(self.supported_pairs, src.suffix, dst.suffix)
        overwrite = bool(kwargs.get("overwrite", False))

        cfg = self._resolve_cfg()
        ensure_dir(dst.parent)
        final_dst = dst if overwrite else unique_output_path(dst)

        table = self._recognize_table(str(src), cfg)
        wb = Workbook()
        try:
            wb.remove(wb.active)
            sheet_name = _safe_sheet_name(src.stem)
            ws = wb.create_sheet(title=sheet_name)
            self._render_table(ws, table)
            wb.save(final_dst)
        finally:
            pass  # openpyxl 无 close

        _log.info("Qwen-VL OCR 完成: %s -> %s", src, final_dst)
        return final_dst

    def convert_many(
        self,
        sources: Iterable[PathLike],
        target: PathLike,
        **kwargs: Any,
    ) -> Path:
        src_list: List[Path] = []
        for item in sources:
            p = Path(item).expanduser()
            if not p.exists():
                raise ConversionError(f"源文件不存在: {p}")
            src_list.append(p)
        if not src_list:
            raise ConversionError("convert_many 至少需要一张图片")

        dst = Path(target).expanduser()
        if not dst.suffix:
            raise ConversionError(f"目标文件必须包含扩展名 .xlsx: {dst}")
        self._check_pair_supported(self.supported_pairs, src_list[0].suffix, dst.suffix)
        overwrite = bool(kwargs.get("overwrite", False))

        cfg = self._resolve_cfg()
        ensure_dir(dst.parent)
        final_dst = dst if overwrite else unique_output_path(dst)

        wb = Workbook()
        try:
            wb.remove(wb.active)
            used_names: set[str] = set()
            for src in src_list:
                ws = wb.create_sheet(title=_safe_sheet_name(src.stem, used=used_names))
                try:
                    table = self._recognize_table(str(src), cfg)
                    self._render_table(ws, table)
                except (ConversionError, QwenVlError) as exc:
                    _log.warning("单张图片 Qwen-VL OCR 失败，留空 sheet (%s): %s", src, exc)
            wb.save(final_dst)
        except OSError as exc:
            raise ConversionError(f"写入 xlsx 失败: {final_dst} ({exc})") from exc

        _log.info("Qwen-VL OCR 批量完成: %d 张 -> %s", len(src_list), final_dst)
        return final_dst

    # ------------------------------------------------------------------ #
    # 内部
    # ------------------------------------------------------------------ #
    def _resolve_cfg(self) -> QwenVlConfig:
        if self.cfg is not None:
            return self.cfg
        # 延迟从 app.config 读取，避免 doc_converter 包硬依赖 FastAPI 项目
        try:
            from app.config import get_settings  # type: ignore[import-not-found]

            s = get_settings()
            api_key = s.qwen_api_key
            if not api_key:
                raise ConversionError(
                    "未配置 QWEN_API_KEY，无法使用 Qwen-VL OCR。"
                    "请在 api/.env 中设置 QWEN_API_KEY=sk-xxx，"
                    "或改用本地 OcrConverter。"
                )
            return QwenVlConfig(
                api_key=api_key,
                base_url=s.qwen_base_url,
                model=s.qwen_model,
                timeout=s.qwen_timeout,
            )
        except ImportError:
            raise ConversionError(
                "未注入 QwenVlConfig 且无法读取 app.config；"
                "请显式传入 cfg= 参数，或在 api 项目内调用。"
            ) from None

    def _recognize_table(self, image_path: str, cfg: QwenVlConfig) -> dict:
        """调用 Qwen-VL 拿回结构化表格 JSON。

        容错策略：has_table=false 时直接抛错（不是表格就别走 OCR）。
        """
        try:
            result = chat_with_image(cfg=cfg, image_path=image_path)
        except QwenVlError as exc:
            raise ConversionError(f"Qwen-VL 识别失败: {exc}") from exc

        if not isinstance(result, dict):
            raise ConversionError(f"Qwen-VL 返回不是 dict: {type(result).__name__}")
        if not result.get("has_table"):
            raise ConversionError(
                "Qwen-VL 判定图片中无表格（has_table=false）。"
                "若图片确实有表格，可尝试切换更优模型或放大图片。"
            )

        header = result.get("header") or []
        rows = result.get("rows") or []
        merges = result.get("merges") or []

        # 基础合法性校验
        if not isinstance(header, list) or not all(isinstance(r, list) for r in header):
            raise ConversionError("Qwen-VL 返回 header 不是二维数组")
        if not isinstance(rows, list) or not all(isinstance(r, list) for r in rows):
            raise ConversionError("Qwen-VL 返回 rows 不是二维数组")
        if not isinstance(merges, list):
            raise ConversionError("Qwen-VL 返回 merges 不是数组")

        return {
            "header": [[_norm_cell(c) for c in row] for row in header],
            "rows": [[_norm_cell(c) for c in row] for row in rows],
            "merges": [m for m in merges if isinstance(m, dict)],
        }

    @staticmethod
    def _render_table(ws, table: dict) -> None:
        """把 JSON 表格渲染到 openpyxl 工作表，保留合并单元格。

        顺序：
        1) 先处理 ``merges`` 字段，标记哪些格子属于合并区（不写值）。
        2) 写非合并区的数据格。
        3) 写合并区 anchor 格（顶部-左侧）的 value。
        4) 设列宽。
        """
        header: List[List[str]] = table["header"]
        rows: List[List[str]] = table["rows"]
        merges: List[dict] = table["merges"]

        # 0) 规范化 merges 为 (r1,c1,r2,c2, value)，丢弃非法项
        norm_merges: List[Tuple[int, int, int, int, str | None]] = []
        for m in merges:
            try:
                r1 = int(m["r1"]); c1 = int(m["c1"])
                r2 = int(m["r2"]); c2 = int(m["c2"])
            except (KeyError, TypeError, ValueError):
                _log.warning("跳过非法 merge: %s", m)
                continue
            if r2 < r1 or c2 < c1:
                _log.warning("跳过反向 merge: %s", m)
                continue
            norm_merges.append((r1, c1, r2, c2, _norm_cell(m.get("value"))))

        # 1) 先建立“合并覆盖集合”并真正合并单元格
        # openpyxl 规则：合并后区域内非 anchor 格变成 MergedCell，value 不可写。
        merge_region: set[Tuple[int, int]] = set()
        for r1, c1, r2, c2, _v in norm_merges:
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    merge_region.add((rr, cc))
            ws.merge_cells(
                start_row=r1 + 1, start_column=c1 + 1,
                end_row=r2 + 1, end_column=c2 + 1,
            )

        # 2) 写表头与数据：跳过合并区所有格子（避免触碰 MergedCell）
        all_rows: List[List[str]] = []
        all_rows.extend(header)
        all_rows.extend(rows)
        n_cols = max((len(r) for r in all_rows), default=0)

        for r_idx0, row in enumerate(all_rows):  # 0-based
            for c_idx0 in range(len(row)):
                if (r_idx0, c_idx0) in merge_region:
                    continue  # 留给 merge 阶段统一写
                val = _norm_cell(row[c_idx0])
                if val is None:
                    continue
                ws.cell(row=r_idx0 + 1, column=c_idx0 + 1, value=val)

        # 3) 写合并区 anchor 单元格的 value（仅 anchor 可写）
        for r1, c1, r2, c2, v in norm_merges:
            if v is None:
                continue
            ws.cell(row=r1 + 1, column=c1 + 1, value=v)

        # 4) 列宽
        for c_idx in range(1, max(n_cols, 1) + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 14


# ---------------------------------------------------------------------- #
# 辅助函数
# ---------------------------------------------------------------------- #
def _norm_cell(v: Any) -> str | None:
    """把模型返回的格子归一化：None / 空字符串统一返回 None。"""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    # 数字 / 布尔直接转字符串，避免被 openpyxl 当作数字处理时格式错乱
    return str(v)


def _safe_sheet_name(name: str, used: set[str] | None = None) -> str:
    """生成合法的 Excel sheet 名（最长 31 字符，去掉非法字符，避免冲突）。"""
    illegal = set("[]:*?/\\")
    cleaned = "".join("_" if ch in illegal else ch for ch in name).strip()
    if not cleaned:
        cleaned = "sheet"
    cleaned = cleaned[:31] or "sheet"
    if used is None:
        return cleaned
    original = cleaned
    i = 2
    while cleaned in used:
        suffix = f"_{i}"
        cleaned = (original[: 31 - len(suffix)] + suffix) if len(original) + len(suffix) > 31 else original + suffix
        i += 1
    used.add(cleaned)
    return cleaned


# 模块加载时自动注册到全局路由
Registry.register(QwenOcrConverter())