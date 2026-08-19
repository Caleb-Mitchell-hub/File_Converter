"""ExcelConverter
~~~~~~~~~~~~~~~~

负责 Excel 文档的"双向"与"出图"转换：

* ``.xlsx`` / ``.xls``  -> ``.pdf``     保留原格式（字体、颜色、边框、合并、图表等）
* ``.pdf``               -> ``.xlsx``    把 PDF 里的表格逐页提取为多个 Sheet
* ``.xlsx`` / ``.ls``    -> ``.png`` / ``.jpg``  高清栅格化（默认 300 DPI）

引擎选择策略
============

Excel -> PDF / 图片：
    * **Windows + 装了 Microsoft Office**：优先使用 :mod:`win32com` 直接调用
      本地 Excel 的 ``ExportAsFixedFormat``，保真度最高（字体、颜色、合并、
      图表全部保留）。每个 Sheet 渲染成 PDF 的一页。
    * **Linux / macOS 或没装 Office**：降级到 LibreOffice headless::

          soffice --headless --convert-to pdf --outdir <dir> <src>

PDF -> Xlsx：
    * 始终使用 :mod:`pdfplumber` 提取 ``page.extract_tables()``，再用
      :mod:`openpyxl` 写入 ``.xlsx``，每页每个 table 单独 Sheet。

Excel -> 图片：
    * 同 Excel -> PDF，先生成中间 PDF，再用 :mod:`fitz` (PyMuPDF) 渲染到
      300 DPI 的位图。
    * 如果目标扩展名是 ``.jpg``，用 :mod:`PIL` 把 PNG 转 JPG（白色填充 alpha）。

依赖库
======

==============  ==============================  ======================
库              用途                            必需?
==============  ==============================  ======================
``openpyxl``    写 .xlsx                         是
``pdfplumber``  解析 PDF 表格                    是
``PyMuPDF``     PDF -> 图片                      是
``Pillow``      PNG -> JPG / alpha 填充          是（仅 JPG 输出时）
``pywin32``     调用本地 Office（最佳保真）      仅 Windows
``subprocess``  调用 LibreOffice headless         是（系统路径需可访问）
==============  ==============================  ======================

平台说明
========

* **Windows + Microsoft Office**：默认走 pywin32，零失真。
* **Linux / macOS**：必须安装 LibreOffice (``soffice`` 命令)。
* **CI / 无 GUI 环境**：只支持 LibreOffice 路径。
"""

from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, ClassVar, Tuple

from ..core.base import BaseConverter, ConversionError, PathLike
from ..core.logger import get_logger
from ..utils.paths import ensure_dir, unique_output_path
from ..utils.platform import OfficeStatus, office_status

# --------------------------------------------------------------------------- #
# 平台相关可选依赖
# --------------------------------------------------------------------------- #
# pywin32 只在 Windows 上可用；其它平台直接 import 会抛 ImportError。
try:  # pragma: no cover - 平台分支
    import win32com.client  # type: ignore[import-not-found]
    _HAS_WIN32 = True
except Exception:  # ImportError 或 pywin32 未安装
    win32com = None  # type: ignore[assignment]
    _HAS_WIN32 = False

# Excel 文件解析 / 写入
try:
    import openpyxl  # type: ignore[import-not-found]
    _HAS_OPENPYXL = True
except Exception:
    openpyxl = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False

# PDF 表格提取
try:
    import pdfplumber  # type: ignore[import-not-found]
    _HAS_PDFPLUMBER = True
except Exception:
    pdfplumber = None  # type: ignore[assignment]
    _HAS_PDFPLUMBER = False

# PDF -> 位图
try:
    import fitz  # type: ignore[import-not-found]  # PyMuPDF
    _HAS_FITZ = True
except Exception:
    fitz = None  # type: ignore[assignment]
    _HAS_FITZ = False

# PNG -> JPG
try:
    from PIL import Image  # type: ignore[import-not-found]
    _HAS_PIL = True
except Exception:
    Image = None  # type: ignore[assignment]
    _HAS_PIL = False

# 显式标出未使用的符号（避免 W0611）
_ = (win32com, openpyxl, pdfplumber, fitz, Image)

# --------------------------------------------------------------------------- #
# 公开类
# --------------------------------------------------------------------------- #
_log = get_logger("converters.excel_converter")
# 模块加载时记录一次平台信息，便于排错
_log.debug("Excel 引擎状态: %s", office_status().preferred_engine())


class ExcelConverter(BaseConverter):
    """Excel 转换器。

    支持的转换组合：

    * ``.xlsx`` -> ``.pdf`` / ``.png`` / ``.jpg``
    * ``.xls``  -> ``.pdf`` / ``.png`` / ``.jpg``
    * ``.pdf``  -> ``.xlsx``

    用法::

        conv = ExcelConverter()
        out = conv.convert("input.xlsx", "output.pdf")
    """

    #: 转换器名称
    name: ClassVar[str] = "ExcelConverter"

    #: 声明支持的扩展名组合（小写、含点）
    supported_pairs: ClassVar[Tuple[Tuple[str, str], ...]] = (
        (".xlsx", ".pdf"),
        (".xls", ".pdf"),
        (".pdf", ".xlsx"),
        (".xlsx", ".png"),
        (".xlsx", ".jpg"),
        (".xls", ".png"),
        (".xls", ".jpg"),
    )

    # ------------------------------------------------------------------ #
    # 构造
    # ------------------------------------------------------------------ #
    def __init__(
        self,
        *,
        dpi: int = 300,
        image_format: str = "png",
        libreoffice_timeout: int = 120,
        status: OfficeStatus | None = None,
        **kwargs: Any,
    ) -> None:
        """构造一个 ExcelConverter 实例。

        Args:
            dpi: 栅格化输出（PNG / JPG）时的目标分辨率，默认 300。
            image_format: 默认的图片格式（``"png"`` 或 ``"jpg"``），会被
                实际目标扩展名覆盖。
            libreoffice_timeout: 调用 ``soffice`` 的最大等待秒数。
            status: 注入的 :class:`OfficeStatus`，主要用于测试。
            **kwargs: 预留扩展位，方便未来增加参数而不破坏签名。
        """
        super().__init__(**kwargs)  # type: ignore[call-arg]
        if dpi < 72:
            raise ValueError("dpi 不应低于 72，否则图像质量过差")
        self.dpi = dpi
        self.image_format = image_format.lower()
        self.libreoffice_timeout = libreoffice_timeout
        self._status = status or office_status()
        _log.debug(
            "ExcelConverter 初始化: dpi=%d, status=%s",
            self.dpi,
            self._status.preferred_engine(),
        )

    # ================================================================== #
    # 公开 API
    # ================================================================== #
    def convert(self, source: PathLike, target: PathLike, **kwargs: Any) -> Path:
        """根据扩展名路由到具体子方法。

        Args:
            source: 源文件路径。
            target: 输出文件路径。
            **kwargs: 透传给子方法的扩展参数。

        Returns:
            实际写入的目标路径（可能因为防覆盖而变成 ``<stem>_1.<ext>``）。
        """
        src, dst = self._resolve_paths(source, target)
        src_ext = src.suffix.lower()
        dst_ext = dst.suffix.lower()
        _log.info("Excel 转换开始: %s -> %s", src, dst)

        # 防覆盖：交给通用工具处理
        final_dst = unique_output_path(dst)
        if final_dst != dst:
            _log.warning("目标已存在，改写为: %s", final_dst)

        # 父目录自动创建
        ensure_dir(final_dst.parent)

        # 路由
        if src_ext in {".xlsx", ".xls"} and dst_ext == ".pdf":
            return self._convert_to_pdf(src, final_dst, **kwargs)
        if src_ext == ".pdf" and dst_ext == ".xlsx":
            return self._pdf_to_xlsx(src, final_dst, **kwargs)
        if src_ext in {".xlsx", ".xls"} and dst_ext in {".png", ".jpg"}:
            return self._convert_to_image(src, final_dst, **kwargs)

        # 不应走到这里（Registry 路由层已经过滤过）
        raise ConversionError(
            f"ExcelConverter 不支持的转换: {src_ext} -> {dst_ext}"
        )

    # ================================================================== #
    # 私有：xlsx/xls -> pdf
    # ================================================================== #
    def _convert_to_pdf(
        self, src: Path, dst: Path, **kwargs: Any
    ) -> Path:
        """把 Excel 文件转换为 PDF。

        优先级：
            1) Windows + Microsoft Office -> ``win32com`` ``ExportAsFixedFormat``
            2) 其它情况 -> LibreOffice headless
        """
        _log.info("xlsx/xls -> pdf 引擎: %s", self._status.preferred_engine())
        try:
            if self._status.has_office and _HAS_WIN32 and sys.platform.startswith("win"):
                return self._excel_to_pdf_win32(src, dst)
            return self._excel_to_pdf_libreoffice(src, dst)
        except ConversionError:
            raise
        except Exception as exc:  # noqa: BLE001 - 统一包装
            raise ConversionError(f"Excel -> PDF 失败: {exc}") from exc

    # ------------------------------------------------------------------ #
    # pywin32 路径
    # ------------------------------------------------------------------ #
    def _excel_to_pdf_win32(self, src: Path, dst: Path) -> Path:
        """通过本地 Excel (COM) 导出 PDF。"""
        if not _HAS_WIN32:
            raise ConversionError("pywin32 不可用，无法调用本地 Office")

        _log.info("使用 pywin32 导出 PDF: %s -> %s", src, dst)
        # 0 = xlTypePDF
        xl_type_pdf = 0
        excel_app = None
        workbook = None
        try:
            excel_app = win32com.client.DispatchEx("Excel.Application")  # type: ignore[name-defined]
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            workbook = excel_app.Workbooks.Open(str(src), ReadOnly=True)
            workbook.ExportAsFixedFormat(xl_type_pdf, str(dst))
        except Exception as exc:
            raise ConversionError(f"pywin32 导出 PDF 失败: {exc}") from exc
        finally:
            # 严格清理 COM 资源
            try:
                if workbook is not None:
                    workbook.Close(False)
            except Exception:  # pragma: no cover
                pass
            try:
                if excel_app is not None:
                    excel_app.Quit()
            except Exception:  # pragma: no cover
                pass

        if not dst.exists():
            raise ConversionError(f"导出完成但目标不存在: {dst}")
        return dst

    # ------------------------------------------------------------------ #
    # LibreOffice 路径
    # ------------------------------------------------------------------ #
    def _excel_to_pdf_libreoffice(self, src: Path, dst: Path) -> Path:
        """通过 LibreOffice headless 把 Excel 转 PDF。"""
        if not self._status.has_libreoffice or not self._status.soffice_path:
            raise ConversionError(
                "未检测到 LibreOffice (`soffice`)。\n"
                + self._status.install_hint()
            )

        soffice = self._status.soffice_path
        # LibreOffice 只能输出到目录，再把生成的文件搬到目标位置
        with tempfile.TemporaryDirectory(prefix="excel2pdf_") as tmp:
            tmp_dir = Path(tmp)
            _log.info("调用 soffice: %s --convert-to pdf", soffice)
            cmd = [
                soffice,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmp_dir),
                str(src),
            ]
            try:
                subprocess.run(
                    cmd,
                    check=True,
                    timeout=self.libreoffice_timeout,
                    capture_output=True,
                )
            except subprocess.TimeoutExpired as exc:
                raise ConversionError(
                    f"soffice 执行超时 (>{self.libreoffice_timeout}s)"
                ) from exc
            except subprocess.CalledProcessError as exc:
                err = (exc.stderr or b"").decode("utf-8", errors="replace")
                raise ConversionError(f"soffice 调用失败: {err}") from exc

            produced = tmp_dir / (src.stem + ".pdf")
            if not produced.exists():
                # 某些 LibreOffice 版本会把点替换成下划线
                alt = tmp_dir / (src.stem.replace(".", "_") + ".pdf")
                if alt.exists():
                    produced = alt
                else:
                    raise ConversionError(
                        f"LibreOffice 未生成 PDF: 期望 {produced}"
                    )
            shutil.move(str(produced), str(dst))
        return dst

    # ================================================================== #
    # 私有：pdf -> xlsx
    # ================================================================== #
    def _pdf_to_xlsx(self, src: Path, dst: Path, **kwargs: Any) -> Path:
        """提取 PDF 中的所有表格为多 Sheet 的 xlsx。"""
        if not _HAS_PDFPLUMBER:
            raise ConversionError("缺少依赖 pdfplumber，请 `pip install pdfplumber`")
        if not _HAS_OPENPYXL:
            raise ConversionError("缺少依赖 openpyxl，请 `pip install openpyxl`")

        _log.info("pdf -> xlsx: %s -> %s", src, dst)
        try:
            with pdfplumber.open(str(src)) as pdf:  # type: ignore[name-defined]
                # openpyxl 必须通过 with 管理
                with openpyxl.Workbook() as wb:  # type: ignore[name-defined]
                    # 默认会有一个空 Sheet，先删掉
                    default_ws = wb.active
                    wb.remove(default_ws)

                    total_tables = 0
                    for page_idx, page in enumerate(pdf.pages, start=1):
                        tables = page.extract_tables() or []
                        for tbl_idx, table in enumerate(tables, start=1):
                            sheet_name = f"P{page_idx}_T{tbl_idx}"[:31]
                            ws = wb.create_sheet(title=sheet_name)
                            for row in table:
                                # 替换 None 为空串，避免 openpyxl 报警
                                ws.append([(c if c is not None else "") for c in row])
                            total_tables += 1
                    _log.info("共写入 %d 个 Sheet", total_tables)
                    wb.save(str(dst))
        except ConversionError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise ConversionError(f"PDF -> XLSX 失败: {exc}") from exc

        if not dst.exists():
            raise ConversionError(f"PDF -> XLSX 完成但目标不存在: {dst}")
        return dst

    # ================================================================== #
    # 私有：xlsx/xls -> png/jpg
    # ================================================================== #
    def _convert_to_image(
        self, src: Path, dst: Path, **kwargs: Any
    ) -> Path:
        """Excel -> 图片。

        关键步骤：
            1. 在临时副本上设置 ``fitToPage`` 打印属性（不修改原文件），
               让 LibreOffice 把整张表缩放到一页宽输出为单页 PDF。
               这样表格是连贯的，不会出现"两页拼起来中间有缝"的情况。
            2. PyMuPDF 把单页 PDF 渲染为指定 DPI 的位图。
        """
        _log.info("xlsx/xls -> image: %s -> %s", src, dst)
        if not _HAS_FITZ:
            raise ConversionError("缺少依赖 PyMuPDF，请 `pip install pymupdf`")
        if not _HAS_PIL:
            raise ConversionError("缺少依赖 Pillow，请 `pip install Pillow`")

        dst_ext = dst.suffix.lower()
        if dst_ext not in {".png", ".jpg"}:
            raise ConversionError(f"目标扩展名必须是 .png 或 .jpg，得到 {dst_ext}")

        # 1) Excel -> 中间 PDF（在临时副本上设置 fitToPage，不动原文件）
        with tempfile.TemporaryDirectory(prefix="excel2img_") as tmp:
            tmp_dir = Path(tmp)
            fit_src = self._prepare_fit_src(src, tmp_dir)
            tmp_pdf = tmp_dir / (src.stem + ".pdf")
            self._convert_to_pdf(fit_src, tmp_pdf)

            # 2) PDF -> 位图（单页：fitToPage 保证了）
            zoom = self.dpi / 72.0
            matrix = fitz.Matrix(zoom, zoom)  # type: ignore[name-defined]
            with fitz.open(str(tmp_pdf)) as doc:  # type: ignore[name-defined]
                if doc.page_count == 0:
                    raise ConversionError(f"中间 PDF 为空: {tmp_pdf}")
                if doc.page_count > 1:
                    _log.warning(
                        "中间 PDF 仍为 %d 页（fitToPage 未生效），将渲染首页",
                        doc.page_count,
                    )
                page = doc.load_page(0)
                px = page.get_pixmap(matrix=matrix, alpha=True)
                img = Image.frombytes(  # type: ignore[name-defined]
                    "RGBA", (px.width, px.height), px.samples
                )

            # 3) 输出
            if dst_ext == ".png":
                img.save(str(dst))
            else:  # .jpg：白底覆盖 alpha
                background = Image.new("RGB", img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[3])
                background.save(str(dst), format="JPEG", quality=95, optimize=True)

        if not dst.exists():
            raise ConversionError(f"Excel -> 图片 完成但目标不存在: {dst}")
        return dst

    # ------------------------------------------------------------------ #
    # 准备 fitToPage 临时副本
    # ------------------------------------------------------------------ #
    def _prepare_fit_src(self, src: Path, tmp_dir: Path) -> Path:
        """复制源 xlsx 到临时目录并设置 ``fitToPage`` 打印属性。

        为什么需要这一步？
        - LibreOffice 默认按 A4 页面分页，宽表会被拆成多页 PDF。
        - 设置 ``fitToWidth=1, fitToHeight=0`` 告诉 LibreOffice：
          "宽度缩放到一页，高度不限"，整张表就只占一页宽。
        - 修改的是临时副本，**不会污染用户原文件**。
        - 如果源是 .xls（openpyxl 不支持），则直接返回原路径，按
          原样让 LibreOffice 拆分（保持向后兼容）。
        """
        if not _HAS_OPENPYXL or src.suffix.lower() != ".xlsx":
            # openpyxl 只支持 .xlsx；.xls 直接走 LibreOffice 默认行为
            return src

        try:
            wb = openpyxl.load_workbook(str(src))  # type: ignore[name-defined]
            for ws in wb.worksheets:
                # 关键三行：让打印时把内容缩放到一页宽
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0  # 0 = 高度不限（多页也行）
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                # 让缩放生效需要把缩放系数清掉（fitToPage 优先于 scale）
                ws.page_setup.scale = None
            fit_src = tmp_dir / (src.stem + "_fit.xlsx")
            wb.save(str(fit_src))
            _log.debug("已生成 fitToPage 临时副本: %s", fit_src)
            return fit_src
        except Exception as exc:  # noqa: BLE001
            # 出错时回退到原文件，不阻塞用户
            _log.warning("准备 fitToPage 副本失败，使用原文件: %s", exc)
            return src

    # ------------------------------------------------------------------ #
    # Pixmap -> JPG (Pillow)
    # ------------------------------------------------------------------ #
    @staticmethod
    def _pixmap_to_jpg(pix: Any, dst: Path) -> None:
        """把 fitz.Pixmap (RGBA) 转成 JPG，白色填充 alpha。"""
        if not _HAS_PIL or Image is None:
            raise ConversionError("缺少 Pillow，无法输出 JPG")
        # pix.samples 是 bytes, n = 通道数 (4 = RGBA)
        n = pix.n
        mode = "RGBA" if n == 4 else "RGB"
        img = Image.frombytes(mode, (pix.width, pix.height), pix.samples)
        if mode == "RGBA":
            background = Image.new("RGB", img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])  # alpha 通道
            img = background
        img.save(str(dst), format="JPEG", quality=95, optimize=True)


# --------------------------------------------------------------------------- #
# 模块加载时自动注册
# --------------------------------------------------------------------------- #
from ..core.registry import Registry  # noqa: E402  -  故意放最后避免循环

Registry.register(ExcelConverter())
