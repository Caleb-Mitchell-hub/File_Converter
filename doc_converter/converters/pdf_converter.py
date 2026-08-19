"""PDF 转换器。

本模块实现 :class:`PdfConverter`，负责把 PDF 文件转换为图片（PNG / JPG）
或提取表格/文本到 Excel 工作簿（XLSX）。

支持的目标格式：
    - ``.pdf`` -> ``.png``  使用 PyMuPDF (fitz) 高保真渲染（默认 300 DPI）
    - ``.pdf`` -> ``.jpg``  使用 PyMuPDF 渲染 + Pillow 保存为 JPEG
    - ``.pdf`` -> ``.xlsx`` 使用 pdfplumber 抽取表格，openpyxl 写入 Excel

多页 PDF 在导出图片时：
    - 1 页：直接写到 ``target``（即使扩展名是 ``.png``/``.jpg`` 也只生成 1 个文件）
    - 多页：把 ``target`` 视为"输出目录 + 基础名"，按 ``page_NNN.png`` 命名

每个 PDF 页面在导出 Excel 时：
    - 该页有表格：每个表格生成一个 sheet，命名 ``p{page}_t{table}``
    - 该页无表格：生成一个文本 sheet，命名 ``p{page}_text``
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any, ClassVar, Tuple

import fitz  # PyMuPDF
import pdfplumber
from openpyxl import Workbook
from PIL import Image

from ..core.base import BaseConverter, ConversionError, PathLike
from ..core.logger import get_logger
from ..core.registry import Registry
from ..utils.paths import ensure_dir, unique_output_path


class PdfConverter(BaseConverter):
    """PDF -> 图片 / Excel 转换器。

    Attributes:
        name: 转换器名称，固定为 ``"PdfConverter"``。
        supported_pairs: 支持的 (源扩展名, 目标扩展名) 组合。
    """

    name: ClassVar[str] = "PdfConverter"
    supported_pairs: ClassVar[Tuple[Tuple[str, str], ...]] = (
        (".pdf", ".png"),
        (".pdf", ".jpg"),
        (".pdf", ".xlsx"),
    )

    _log = get_logger("converters.pdf")

    # ------------------------------------------------------------------ #
    # 构造
    # ------------------------------------------------------------------ #
    def __init__(
        self,
        *,
        dpi: int = 300,
        jpg_quality: int = 95,
        **kwargs: Any,
    ) -> None:
        """初始化 PdfConverter。

        Args:
            dpi: 渲染图片时的 DPI，默认 300（满足"高清"要求）。
            jpg_quality: JPEG 编码质量，1-95+，默认 95。
            **kwargs: 透传给父类的扩展参数（保留扩展性）。
        """
        super().__init__(**kwargs)  # type: ignore[call-arg]
        if dpi <= 0:
            raise ConversionError(f"dpi 必须为正整数，收到: {dpi}")
        if not (1 <= jpg_quality <= 100):
            raise ConversionError(f"jpg_quality 必须在 1-100 之间，收到: {jpg_quality}")
        self.dpi = dpi
        self.jpg_quality = jpg_quality

    # ------------------------------------------------------------------ #
    # 公共入口
    # ------------------------------------------------------------------ #
    def convert(self, source: PathLike, target: PathLike, **kwargs: Any) -> Path:
        """根据 ``target`` 扩展名分发到具体的转换方法。

        Args:
            source: 源 PDF 文件路径。
            target: 输出文件路径（或多页 PNG 时的"输出目录 + 基础名"）。
            **kwargs: 可覆盖实例参数（``dpi`` / ``jpg_quality``）。

        Returns:
            实际写入的路径。单页图片时即 ``target``；
            多页图片时返回"第一个输出文件"的路径。
        """
        # 允许通过 kwargs 临时覆盖渲染参数
        if "dpi" in kwargs:
            self.dpi = int(kwargs["dpi"])
        if "jpg_quality" in kwargs:
            self.jpg_quality = int(kwargs["jpg_quality"])

        src, dst = self._resolve_paths(source, target)
        dst_ext = dst.suffix.lower()

        if dst_ext == ".png":
            return self._to_image(src, dst, fmt="png")
        if dst_ext == ".jpg" or dst_ext == ".jpeg":
            return self._to_image(src, dst, fmt="jpg")
        if dst_ext == ".xlsx":
            return self._to_xlsx(src, dst)

        raise ConversionError(f"不支持的目标扩展名: {dst_ext}")

    # ------------------------------------------------------------------ #
    # PDF -> PNG / JPG
    # ------------------------------------------------------------------ #
    def _to_image(
        self,
        src: Path,
        dst: Path,
        *,
        fmt: str,
    ) -> Path:
        """把 PDF 渲染为 PNG 或 JPG 图片。

        Args:
            src: 源 PDF。
            dst: 目标文件路径（多页时视为"目录 + 基础名"）。
            fmt: ``"png"`` 或 ``"jpg"``。

        Returns:
            实际写入的第一个文件路径（多页场景下后续文件与该路径同目录）。
        """
        self._log.info("PDF -> %s 开始: %s -> %s (dpi=%d)", fmt.upper(), src, dst, self.dpi)
        try:
            with fitz.open(src) as doc:
                page_count = doc.page_count
                if page_count == 0:
                    raise ConversionError(f"PDF 没有任何页面: {src}")

                zoom = self.dpi / 72.0
                mat = fitz.Matrix(zoom, zoom)

                if page_count == 1:
                    # 单页：直接写到 dst
                    out_path = unique_output_path(dst)
                    ensure_dir(out_path.parent)
                    page = doc.load_page(0)
                    self._render_page(page, mat, fmt, out_path)
                    self._log.info("PDF -> %s 完成: %s", fmt.upper(), out_path)
                    return out_path

                # 多页：把 dst 视为"目录 + 基础名"
                # 1) 用户可能传入的是 "out_dir/out.png"，多页时把 out.png 视为基础名
                out_dir = dst if dst.is_dir() else dst.parent
                ensure_dir(out_dir)
                base = dst.stem  # 不含扩展名，作为前缀
                ext = ".png" if fmt == "png" else ".jpg"

                first_path: Path | None = None
                for i in range(page_count):
                    page = doc.load_page(i)
                    page_path = out_dir / f"{base}_page_{i + 1:03d}{ext}"
                    page_path = unique_output_path(page_path)
                    self._render_page(page, mat, fmt, page_path)
                    if first_path is None:
                        first_path = page_path
                self._log.info(
                    "PDF -> %s 完成: %d 页 -> %s",
                    fmt.upper(),
                    page_count,
                    out_dir,
                )
                if first_path is None:  # pragma: no cover - 防御
                    raise ConversionError("未生成任何输出文件")
                return first_path
        except ConversionError:
            raise
        except fitz.FileDataError as exc:
            raise ConversionError(f"PDF 文件损坏或无法解析: {src}") from exc
        except fitz.FitzError as exc:  # type: ignore[attr-defined]
            raise ConversionError(f"PyMuPDF 错误: {exc}") from exc
        except Exception as exc:  # noqa: BLE001
            raise ConversionError(f"PDF -> {fmt.upper()} 失败: {exc}") from exc

    def _render_page(
        self,
        page: "fitz.Page",
        mat: "fitz.Matrix",
        fmt: str,
        out_path: Path,
    ) -> None:
        """渲染单页并写入文件。"""
        if fmt == "png":
            pix = page.get_pixmap(matrix=mat, alpha=False)
            pix.save(str(out_path))
        elif fmt == "jpg":
            # JPG 不支持 alpha，因此显式指定 RGB 色彩空间
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB, alpha=False)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=self.jpg_quality, optimize=True)
            out_path.write_bytes(buf.getvalue())
        else:
            raise ConversionError(f"内部错误：未知 fmt: {fmt}")

    # ------------------------------------------------------------------ #
    # PDF -> XLSX
    # ------------------------------------------------------------------ #
    def _to_xlsx(self, src: Path, dst: Path) -> Path:
        """把 PDF 抽取为 Excel 工作簿。

        每个 PDF 页面：
            - 若包含表格：每个表格生成一个 sheet，命名 ``p{page}_t{idx}``
            - 若没有表格：生成一个文本 sheet，命名 ``p{page}_text``

        Args:
            src: 源 PDF。
            dst: 目标 .xlsx 路径。

        Returns:
            实际写入的 xlsx 路径（可能因 ``unique_output_path`` 后缀不同）。
        """
        self._log.info("PDF -> XLSX 开始: %s -> %s", src, dst)
        try:
            out_path = unique_output_path(dst)
            ensure_dir(out_path.parent)

            wb = Workbook()
            try:
                # Workbook() 默认带一个 active sheet，删除以避免空 sheet
                if wb.active is not None:
                    wb.remove(wb.active)

                with pdfplumber.open(src) as pdf:
                    total_pages = len(pdf.pages)
                    if total_pages == 0:
                        raise ConversionError(f"PDF 没有任何页面: {src}")

                    for page_idx, page in enumerate(pdf.pages, start=1):
                        tables = page.extract_tables() or []
                        if tables:
                            for t_idx, table in enumerate(tables, start=1):
                                sheet_name = self._safe_sheet_name(
                                    f"p{page_idx}_t{t_idx}"
                                )
                                ws = wb.create_sheet(sheet_name)
                                for row in table:
                                    # 单元格可能为 None，统一转为空字符串
                                    ws.append(
                                        [(c if c is not None else "") for c in row]
                                    )
                        else:
                            # 没有表格：把整页文本按行写入
                            sheet_name = self._safe_sheet_name(
                                f"p{page_idx}_text"
                            )
                            ws = wb.create_sheet(sheet_name)
                            text = page.extract_text() or ""
                            for line in text.splitlines():
                                ws.append([line])

                # 防止一个 sheet 都没有的极端情况
                if not wb.sheetnames:
                    wb.create_sheet("empty").append(["(空文档)"])

                wb.save(out_path)
            finally:
                # openpyxl Workbook 不支持 with 上下文，手动关闭
                wb.close()
            self._log.info("PDF -> XLSX 完成: %s", out_path)
            return out_path
        except ConversionError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise ConversionError(f"PDF -> XLSX 失败: {exc}") from exc

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """生成 openpyxl 合法的 sheet 名（<= 31 字符，且不含非法字符）。"""
        illegal = set("[]:*?/\\")
        cleaned = "".join("_" if ch in illegal else ch for ch in name)
        return cleaned[:31] or "Sheet"


# ---------------------------------------------------------------------- #
# 模块加载时自动注册到全局路由
# ---------------------------------------------------------------------- #
Registry.register(PdfConverter())
